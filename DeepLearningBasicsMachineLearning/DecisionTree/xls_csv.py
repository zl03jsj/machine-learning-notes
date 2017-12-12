# -*- coding: utf-8 -*-
import xlrd
import csv

# from datetime import date, datetime
from openpyxl import load_workbook
# from openpyxl import Workbook
from openpyxl.compat import range


# from openpyxl.cell import get_column_letter


def read_excel(filename):
    workbook = xlrd.open_workbook(filename)
    # print sheet2.name,sheet2.nrows,sheet2.ncols
    sheet2 = workbook.sheet_by_index(0)

    for row in range(0, sheet2.nrows):
        rows = sheet2.row_values(row)

        def _tostr(cell):
            if type(u'') == type(cell):
                return "\"%s\"" % cell.encode('utf8')
            else:
                return "\"%s\"" % str(cell)

    print(','.join([_tostr(cell) for cell in rows]))


def excel_to_csv(xls_filename, sheet_name, csv_filename):
    try:
        xlsx_file_reader = load_workbook(filename=xls_filename)
        for sheet in xlsx_file_reader.get_sheet_names():
            # # 每个sheet输出到一个csv文件中，文件名用xlsx文件名和sheet名用'_'连接
            # csv_filename = '{xlsx}_{sheet}.csv'.format(
            #     xlsx=os.path.splitext(xls_filename.replace(' ', '_'))[0],
            #     sheet=sheet.replace(' ', '_'))
            if sheet_name != sheet: continue

            csv_file = open(csv_filename, 'w+')
            csv_file_writer = csv.writer(csv_file)

            sheet_ranges = xlsx_file_reader[sheet]
            for row in sheet_ranges.rows:
                row_container = []
                for cell in row:
                    if type(cell.value) == 'unicode':
                        row_container.append(cell.value.encode('utf-8'))
                    else:
                        row_container.append(str(cell.value))
                csv_file_writer.writerow(row_container)
            csv_file.close()

    except Exception as e:
        print(e)


        

# if __name__ == '__main__':
#     if len(sys.argv) != 2:
#         print('usage: xlsx2csv <xlsx file name>')
#     else:
#         xlsx2csv(sys.argv[1])
#     sys.exit(0)

# def xls_to_cvs(filename):
#     try:
#         #open the file you want to process
#         xlsx_file_reader = load_workbook(filename)
#         #every sheet output to a csv file，filename is xlsx filename and sheet filename  '_' connection
#         for sheet in xlsx_file_reader.get_sheet_names():
#
#             #craete file with file name
#             csv_filename = '{xlsx}.csv'.format(xlsx = os.path.splitext(filename.replace(' ', '_'))[0])
#
#             #write csv file
#             csv_file = file(csv_filename, 'wb')
#             csv_file_writer = csv.writer(csv_file)
#
#             #read the sheets of excel file
#             sheet_ranges = xlsx_file_reader[sheet]
#
#             #Loop traversal every sheets' data
#             for row in sheet_ranges.rows:
#                 row_container = []
#                 for cell in row:
#                     if type(cell.value) == 'unicode':
#                         row_container.append(cell.value.encode('utf-8'))
#                     else:
#                         row_container.append(str(cell.value))
#                 csv_file_writer.writerow(row_container)
#             csv_file.close()
#     except Exception as e:
#         print e
#     return csv_filename
